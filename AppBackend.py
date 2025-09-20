from flask import Flask, request, render_template
from twilio.rest import Client
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)

# Twilio credentials (add yours here)
account_sid = ""
auth_token = ""
twilio_phone_number = ""

client = Client(account_sid, auth_token)

EXCEL_FILE = "appointments.xlsx"

def save_to_excel(name, mobile, reason):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Full Name", "Mobile", "Reason", "Timestamp"])  # headers

    # Add row with timestamp
    ws.append([name, mobile, reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(EXCEL_FILE)

@app.route('/')
def home():
    return render_template("Rs_appointments.html")  

@app.route('/submit_appointment', methods=['POST'])
def submit_appointment():
    # Get form data
    name = request.form.get("name")
    mobile = request.form.get("mobile")
    reason = request.form.get("reason")

    # Save to Excel
    save_to_excel(name, mobile, reason)

    # SMS to patient
    sms_body = f"Hello {name}, your appointment request has been received. Reason: {reason}. We will contact you soon. - Clinic"
    client.messages.create(
        body=sms_body,
        from_=twilio_phone_number,
        to=f"+91{mobile}"  # assuming India
    )

    # SMS to clinic staff
    clinic_phone = "+918002147890"  # replace with clinic’s number
    staff_message = f"New appointment request:\nName: {name}\nMobile: {mobile}\nReason: {reason}"
    client.messages.create(
        body=staff_message,
        from_=twilio_phone_number,
        to=clinic_phone
    )

    # Render confirmation page with name + reason
    return render_template("Rs_appointments_confirmation.html", name=name, reason=reason)

if __name__ == "__main__":
    app.run(debug=True)
